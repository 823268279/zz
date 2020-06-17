#coding=utf-8


import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

import datetime
from py._xmlgen import html
import pytest
import requests
import json
import random

# 添加接口地址与项目名称
def pytest_configure(config):
    config._metadata["项目名称"] = "NEW_CRM_v1.0"
    config._metadata['测试地址'] = 'http://api.newcrm.group.weixin.wuerp.com'
# 添加所属部门与测试人员
@pytest.mark.optionalhook
def pytest_html_results_summary(prefix):
    prefix.extend([html.p("所属部门: 测试部")])
    prefix.extend([html.p("测试人员: wowo")])



#获取headers
@pytest.fixture(scope='session')    
def headers():
    url='http://api.newcrm.group.weixin.wuerp.com/api/v1.0/Token/Get'
    data={
        "id":"wx5f1e0495dc540bb1",
        "key":"7947ff3cbf06647f312e4e6ec5e32943"
        }
    try:
        response=requests.post(url=url,data=data)
        response_json=response.json()
        assert response.status_code == 200
        assert response_json['message'] =='获取授权成功'
        headers={}
        headers['Authorization']=response_json['data']['Data']['token']
        return headers
    except:
        raise
    
@pytest.fixture(scope='session')    
def menber():
    data={
        "CpnID":'0001',
        "SubID":'3378049226@qq.com',
        "url":'http://api.newcrm.group.weixin.wuerp.com/member/v1.0%s'
        }
    return data

@pytest.fixture(scope='session')   
def manage():
    data={
        "CpnID":'0001',
        "SubID":'3378049226@qq.com',
        "url":'http://api.newcrm.group.weixin.wuerp.com/manage/v1.0%s'
        }
    return data


#获取当前时间
@pytest.fixture(scope='function')  
def now_time():
    data={}
    data['ymd']=datetime.datetime.now().strftime('%Y-%m-%d')
    data['ymd_hms']=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data['StDt']=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data['EdDt']=(datetime.datetime.now()+datetime.timedelta(days=3)).strftime('%Y-%m-%d %H:%M:%S')
    return data


#随机会员数据
@pytest.fixture(scope='module')   
def menber_data_random():
    data={}
    a = ['01','02','03','04']
    b = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                'U', 'V', 'W', 'X', 'Y', 'Z']
    c = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't',
            'u', 'v', 'w', 'x', 'y', 'z']
    #openid
    x = '%s%s%s%s%s%s%s-'%(random.choice(b),random.choice(b),random.choice(b),random.choice(b),random.choice(b),random.choice(b),random.choice(b))
    y = '%s%s%s-%s%s%s%s'%(random.choice(c),random.choice(c),random.choice(c),random.choice(c),random.choice(c),random.choice(c),random.choice(c))
    z = '%s%s%s%s%s-%s%s'%(random.choice(b),random.choice(b),random.choice(b),random.choice(b),random.choice(b),random.choice(b),random.choice(b))
    data['OpnID']='%s%s%s'%(x,y,z)
    #卡类型
    data['VipTpID']=random.choice(a)  
    #姓名
    q1 = '%s%s' % (random.choice(b),random.choice(b))
    q2 = '%s%s%s%s%s' % (random.choice(c),random.choice(c),random.choice(c),random.choice(c),random.choice(c))
    data['Name']='%s%s' % (q1,q2)
    #手机
    data['Tel']='13%s%s' % (sum(random.sample(range(10000,100000),1)),sum(random.sample(range(1000,10000),1)))
    #生日
    e1 = '19%s'% (sum(random.sample(range(10,100),1)))
    e2 = '0%s'% sum(random.sample(range(1,10),1))
    e3 = '%s'% sum(random.sample(range(10,29),1))
    data['Brth']='%s-%s-%s'% (e1,e2,e3)
    #身份证
    r1 = '%s'% sum(random.sample(range(1000,10000),1)) 
    r2 = '%s%s%s'% (e1,e2,e3)
    r3 = '%s'% sum(random.sample(range(1000,10000),1)) 
    data['IDntNmb']='51%s%s%s'% (r1,r2,r3)
    return data



#随机停车场信息
@pytest.fixture(scope='function')   
def parking_data_random():
    data={}
    #停车场编号
    data['ParkID']=sum(random.sample(range(10000000,999999999),4))
    #故障热线
    data['Tel']='13%s%s' % (sum(random.sample(range(10000,100000),1)),sum(random.sample(range(1000,10000),1)))
    return data


#获取会员手机注册数据
@pytest.fixture(scope='session')    
def menber_register_data():
    __file__ = r'../test_data/2020-05-09-09-27_test_data.xlsx'
    workbook=load_workbook(__file__)
    worksheet=workbook['Sheet1']
    response_menber_register=eval(worksheet.cell(2,6).value,{"true":"0","false":"0"})['data']['Data'][0]
    return response_menber_register

#获取停车场分页
@pytest.fixture(scope='session')    
def parking_page_data():
    __file__ = r'../test_data/2020-05-09-09-27_test_data.xlsx'
    workbook=load_workbook(__file__)
    worksheet=workbook['Sheet1']
    response_parking_page=eval(worksheet.cell(3,6).value,{"true":"0","false":"0"})['data']['PageDataList'][0]
    return response_parking_page

#获取停车场缴费规则分页
@pytest.fixture(scope='session')    
def parking_rule_page_data():
    __file__ = r'../test_data/2020-05-09-09-27_test_data.xlsx'
    workbook=load_workbook(__file__)
    worksheet=workbook['Sheet1']
    response_parking_rule_page=eval(worksheet.cell(4,6).value,{"true":"0","false":"0"})['data']['PageDataList'][0]
    return response_parking_rule_page